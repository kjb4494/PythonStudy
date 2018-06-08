import openpyxl as xl
import pprint
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill, Color
from openpyxl.styles import Border, Side
import random

# https://media.readthedocs.org/pdf/openpyxl/latest/openpyxl.pdf

def excel_ex01():
    # 엑셀파일 읽어 오기
    sheet_workbook = xl.load_workbook("C:\\Users\\kitri\\Desktop\\stu.xlsx")
    print(sheet_workbook)

    # 시트이름 출력하기
    sheet = sheet_workbook[sheet_workbook.sheetnames[0]]
    print(sheet)

    # 셸에 접근하기
    print(sheet['A1'].value)

    listv = [chr(i) for i in range(ord('A'), ord('J') + 1)]
    for i in listv:
        print(sheet[i + str(1)].value, end=' ')

    print()
    print(sheet.cell(row=1, column=1).value)


def excel_q01():
    # 1행 1열부터 10행 10열까지 데이터 가져오기
    sheet_workbook = xl.load_workbook("C:\\Users\\kitri\\Desktop\\stu.xlsx")
    sheet = sheet_workbook[sheet_workbook.sheetnames[0]]
    matrix = [[' '] * 10 for i in range(10)]
    for i in range(10):
        for j in range(10):
            matrix[i][j] = sheet.cell(row=i + 1, column=j + 1).value
    pprint.pprint(matrix)


def excel_ex02():
    class EXCEL_:
        def __init__(self):
            self.wb = Workbook()
            self.ws = self.wb.active

        def __del__(self):
            self.wb.save("C:\\Users\\kitri\\Desktop\\stu.xlsx")

    ojExcel = EXCEL_()

    ojExcel.ws.title = "kitri"
    ojExcel.ws['A1'] = "Hello Wolrd"


def excel_ex03():
    wb = load_workbook(
        "C:\\Users\\kitri\\Desktop\\stu.xlsx"
    )
    wsCreate = wb.create_sheet(title="GoGossing")

    # 시트에 값 적재
    num = 1
    for row in range(1, 11):
        for col in range(1, 11):
            wsCreate.cell(row=row, column=col, value=num * 10)
            num += 1

    wb.save("C:\\Users\\kitri\\Desktop\\stu.xlsx")


# 정렬
def excel_ex04():
    wb = xl.Workbook()
    ws = wb.active
    ws['A1'] = "Go to the home ~~~!"
    cell_value = ws['A1']
    cell_value.font = Font(name='consolas',
                           size=20,
                           bold=True)
    cell_value.fill = PatternFill(patternType='solid',
                                  fgColor=Color('a9d9bc'))
    ws.merge_cells("A1:H1")
    cell_value.alignment = Alignment(horizontal='center',
                                     vertical='center')
    wb.save("C:\\Users\\kitri\\Desktop\\stu.xlsx")

def excel_ex05():
    workBook = xl.Workbook()  # 객체 생성
    workSheet = workBook.active
    workSheet['A2'] = "HI"
    cellV = workSheet['A2']
    # 네모 박스를 설정
    box = Border(left=Side(border_style="thin", color='FF000000'),
                 right=Side(border_style="thin", color='FF000000'),
                 top=Side(border_style="thin", color='FF000000'),
                 bottom=Side(border_style="thin", color='FF000000'),
                 diagonal=Side(border_style="thin", color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style="thin", color='FF000000'),
                 vertical=Side(border_style="thin", color='FF000000'),
                 horizontal=Side(border_style="thin", color='FF000000'))
    cellV.border = box
    workBook.save("C:\\Users\\kitri\\Desktop\\stu.xlsx")

def excel_q02():
    # 로또 파일 만들기
    class LottoExcel:
        def __init__(self):
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "lotto List"
            self.ws['B2'] = "Lottp"
            self.ws['B4'] = "당첨번호"
            self.ws['E4'] = "보너스번호"
            self.ws['B7'] = "랜덤로또"
            self.ws['E7'] = "등수"

            self.sampleList = [i for i in range(1, 46)]

            self.lottoNumList = [0, 0, 0, 0, 0, 0, 0]
            self.selectedNumList = [0, 0, 0, 0, 0, 0, 0]
            self.bonusIndex = random.randint(1, 7)
            self.rank = ''

        def __del__(self):
            self.wb.save("C:\\Users\\kitri\\Desktop\\lotto.xlsx")

        def generateRandomList(self):
            randomList = random.sample(self.sampleList, 7)
            randomList.sort()
            return randomList

        def createLotto(self):
            self.lottoNumList = self.generateRandomList()
            self.ws['B5'] = self.getLottoNumber()
            self.setBonusNumber()

        def getLottoNumber(self):
            return ' '.join([str(s) for s in self.lottoNumList])

        def getSelectedLotto(self):
            self.selectedNumList = self.generateRandomList()
            return ' '.join([str(s) for s in self.selectedNumList])

        def getRank(self):
            if self.luckCount() == 6 and not self.isInBonus():
                self.rank = "1등"
            elif self.luckCount() == 6:
                self.rank = "2등"
            elif self.luckCount() == 5 and not self.isInBonus():
                self.rank = "3등"
            elif self.luckCount() == 4 and not self.isInBonus():
                self.rank = "4등"
            elif self.luckCount() == 3 and not self.isInBonus():
                self.rank = "5등"
            else:
                self.rank = "꽝! 다음기회에 ^^"
            return self.rank

        def luckCount(self):
            return len(set(self.lottoNumList) & set(self.selectedNumList))

        def isInBonus(self):
            return self.lottoNumList[self.bonusIndex] in self.selectedNumList

        def setBonusNumber(self):
            self.ws['E5'] = str(self.lottoNumList[self.bonusIndex])

    ojLottoExcel = LottoExcel()
    ojLottoExcel.createLotto()
    for i in range(100000):
        indexB = 'B' + str(9+i)
        ojLottoExcel.ws[indexB] = ojLottoExcel.getSelectedLotto()
        indexE = 'E' + str(9+i)
        ojLottoExcel.ws[indexE] = ojLottoExcel.getRank()


def excel_output():
    # excel_ex01()
    # excel_q01()
    # excel_ex02()
    # excel_ex03()
    # excel_ex04()
    # excel_ex05()
    excel_q02()